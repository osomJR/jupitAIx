from __future__ import annotations

"""
Production-grade structured extraction engine for jupitAIx v1.

Purpose:
- own request validation, deterministic extraction, artifact generation, and response construction
  for the structured_extract feature outside analyzer.py
- remain aligned with schema.py, validation.py, extraction.py, and Product Contract v1
- support both single-document and document-set inputs
- provide document-class-aware extraction strategies and strict result-shape enforcement

Design principles:
- stateless and deterministic
- source-grounded: every extracted value is derived from the provided document text only
- no external knowledge or enrichment
- explicit human-review requirement before reliance/final export
- clean writer layer for JSON, CSV, and XLSX artifacts
"""

from dataclasses import dataclass, field
from datetime import datetime, timezone
import csv
import json
import re
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Protocol, Sequence, Union

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.extraction import extract_text_by_format
from src.schema import (
    AnalyzerRequest,
    AnalyzerResponse,
    DocumentPayload,
    DocumentSetPayload,
    FeatureType,
    HumanReviewRequirement,
    OutputPolicy,
    StructuredDataOutputFormat,
    StructuredExtractionDocumentClass,
    StructuredExtractionRequest,
    StructuredExtractionResultShape,
)
from src.validation import (
    build_structured_extraction_file_result,
    validate_analyzer_request,
    validate_analyzer_response,
)


STRUCTURED_EXTRACTION_RULES = """
TASK: STRUCTURED EXTRACTION
RULES:
- Extract predefined or user-selected fields strictly from the provided document or document set
- Preserve source meaning without interpretation or invention
- No external knowledge may be introduced
- Output may be key-value fields, tables, row-based records, or machine-readable structured data
- Human review remains required before reliance or final export
""".strip()


# =========================
# Artifact and config models
# =========================


@dataclass(frozen=True)
class PersistedArtifact:
    file_name: str
    file_extension: str
    file_size_mb: float
    file_path: str
    storage_key: Optional[str] = None
    download_url: Optional[str] = None


@dataclass(frozen=True)
class StructuredExtractionConfig:
    algorithm_version: Optional[str] = "structured-extraction-v1.0.0"
    artifact_base_dir: str = "artifacts/structured_extraction"
    include_empty_selected_fields: bool = True
    max_context_excerpt_chars: int = 240


@dataclass(frozen=True)
class SourceEvidence:
    source_document_index: int
    field_name: str
    value: str
    line_number: Optional[int] = None
    excerpt: Optional[str] = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_document_index": self.source_document_index,
            "field_name": self.field_name,
            "value": self.value,
            "line_number": self.line_number,
            "excerpt": self.excerpt,
        }


@dataclass(frozen=True)
class ExtractedField:
    name: str
    value: Any
    source_document_index: int
    confidence: float = 1.0
    evidence: list[SourceEvidence] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "value": self.value,
            "source_document_index": self.source_document_index,
            "confidence": self.confidence,
            "evidence": [item.to_dict() for item in self.evidence],
        }


@dataclass(frozen=True)
class ExtractedTable:
    table_index: int
    source_document_index: int
    columns: list[str]
    rows: list[dict[str, Any]]

    def to_dict(self) -> dict[str, Any]:
        return {
            "table_index": self.table_index,
            "source_document_index": self.source_document_index,
            "columns": list(self.columns),
            "rows": [dict(row) for row in self.rows],
        }


@dataclass(frozen=True)
class DocumentExtraction:
    source_document_index: int
    filename: Optional[str]
    input_format: str
    ocr_used: bool
    document_classes: list[str]
    fields: list[ExtractedField]
    tables: list[ExtractedTable]
    row_records: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)

    def fields_as_mapping(self) -> dict[str, Any]:
        output: dict[str, Any] = {}
        for item in self.fields:
            output[item.name] = item.value
        return output

    def to_machine_readable(self) -> dict[str, Any]:
        return {
            "source_document_index": self.source_document_index,
            "filename": self.filename,
            "input_format": self.input_format,
            "ocr_used": self.ocr_used,
            "document_classes": list(self.document_classes),
            "fields": [field_item.to_dict() for field_item in self.fields],
            "tables": [table.to_dict() for table in self.tables],
            "row_based_records": [dict(row) for row in self.row_records],
            "warnings": list(self.warnings),
        }


@dataclass(frozen=True)
class StructuredExtractionOutput:
    payload: dict[str, Any]
    tabular_rows: list[dict[str, Any]]
    documents: list[DocumentExtraction]


# =========================
# Strategy interfaces
# =========================


class DocumentClassExtractionStrategy(Protocol):
    document_class: StructuredExtractionDocumentClass

    def extract(
        self,
        *,
        text: str,
        lines: Sequence[str],
        source_document_index: int,
        selected_fields: Sequence[str],
        config: StructuredExtractionConfig,
    ) -> tuple[list[ExtractedField], list[dict[str, Any]], list[str]]:
        """Return fields, class-specific row records, warnings."""
        ...


class StructuredExtractionBackend(Protocol):
    def extract(
        self,
        *,
        documents: Sequence[DocumentPayload],
        result_shape: StructuredExtractionResultShape,
        selected_fields: Sequence[str],
        document_classes: Sequence[StructuredExtractionDocumentClass],
        config: StructuredExtractionConfig,
    ) -> StructuredExtractionOutput:
        ...


# =========================
# Deterministic backend
# =========================


class DeterministicStructuredExtractionBackend:
    """
    Rule-based, document-class-aware extraction backend.

    This backend is intentionally deterministic. It does not infer facts outside
    the document text. It extracts explicit key-value pairs, tabular structures,
    and class-specific fields/records using conservative regex patterns.
    """

    def __init__(self, strategies: Optional[Sequence[DocumentClassExtractionStrategy]] = None) -> None:
        provided = list(strategies or _default_strategies())
        self._strategies: dict[StructuredExtractionDocumentClass, DocumentClassExtractionStrategy] = {
            strategy.document_class: strategy for strategy in provided
        }

    def extract(
        self,
        *,
        documents: Sequence[DocumentPayload],
        result_shape: StructuredExtractionResultShape,
        selected_fields: Sequence[str],
        document_classes: Sequence[StructuredExtractionDocumentClass],
        config: StructuredExtractionConfig,
    ) -> StructuredExtractionOutput:
        if not documents:
            raise ValueError("structured_extract requires at least one document.")

        normalized_selected = _normalize_selected_fields(selected_fields)
        resolved_classes = list(document_classes) or [StructuredExtractionDocumentClass.form]

        document_outputs: list[DocumentExtraction] = []
        for index, document in enumerate(documents):
            text = _ensure_document_text(document)
            lines = list(_iter_lines(text))
            generic_fields = _extract_generic_key_values(
                lines=lines,
                source_document_index=index,
                selected_fields=normalized_selected,
                config=config,
            )
            generic_tables = _extract_generic_tables(lines=lines, source_document_index=index)

            class_fields: list[ExtractedField] = []
            class_rows: list[dict[str, Any]] = []
            warnings: list[str] = []

            for document_class in resolved_classes:
                strategy = self._strategies.get(document_class)
                if strategy is None:
                    warnings.append(f"No specialized extractor registered for document_class='{document_class.value}'.")
                    continue
                fields, rows, strategy_warnings = strategy.extract(
                    text=text,
                    lines=lines,
                    source_document_index=index,
                    selected_fields=normalized_selected,
                    config=config,
                )
                class_fields.extend(fields)
                class_rows.extend(rows)
                warnings.extend(strategy_warnings)

            fields = _merge_fields(
                [*generic_fields, *class_fields],
                selected_fields=normalized_selected,
                include_empty_selected_fields=config.include_empty_selected_fields,
                source_document_index=index,
            )
            row_records = _derive_row_records(
                source_document_index=index,
                fields=fields,
                tables=generic_tables,
                class_rows=class_rows,
            )
            document_outputs.append(
                DocumentExtraction(
                    source_document_index=index,
                    filename=document.filename,
                    input_format=document.metadata.input_format.value,
                    ocr_used=document.metadata.ocr_used,
                    document_classes=[item.value for item in resolved_classes],
                    fields=fields,
                    tables=generic_tables,
                    row_records=row_records,
                    warnings=warnings,
                )
            )

        shaped_payload = ShapeEnforcer.enforce(
            result_shape=result_shape,
            selected_fields=list(selected_fields),
            documents=document_outputs,
        )
        tabular_rows = ShapeEnforcer.to_tabular_rows(
            result_shape=result_shape,
            selected_fields=list(selected_fields),
            documents=document_outputs,
        )
        return StructuredExtractionOutput(
            payload=shaped_payload,
            tabular_rows=tabular_rows,
            documents=document_outputs,
        )

def _rx(pattern: str) -> re.Pattern[str]:
    return re.compile(pattern, re.IGNORECASE)


def _money_rx(label_pattern: str) -> re.Pattern[str]:
    money = r"(?:NGN|₦|USD|\$|EUR|€|GBP|£)?\s?[0-9][0-9,]*(?:\.\d{2})?"
    return _rx(rf"\b(?:{label_pattern})\s*[:\-]?\s*(?P<value>{money})\b")

# =========================
# Document-class strategies
# =========================

class BaseRegexStrategy:
    document_class: StructuredExtractionDocumentClass
    field_patterns: Mapping[str, Sequence[re.Pattern[str]]] = {}

    def extract(
        self,
        *,
        text: str,
        lines: Sequence[str],
        source_document_index: int,
        selected_fields: Sequence[str],
        config: StructuredExtractionConfig,
    ) -> tuple[list[ExtractedField], list[dict[str, Any]], list[str]]:
        del text
        fields: list[ExtractedField] = []
        warnings: list[str] = []
        for field_name, patterns in self.field_patterns.items():
            if selected_fields and _normalize_field_name(field_name) not in selected_fields:
                continue
            match = _first_regex_match(lines, patterns)
            if match is None:
                continue
            value, line_number, excerpt = match
            fields.append(
                ExtractedField(
                    name=field_name,
                    value=value,
                    source_document_index=source_document_index,
                    confidence=0.92,
                    evidence=[
                        SourceEvidence(
                            source_document_index=source_document_index,
                            field_name=field_name,
                            value=value,
                            line_number=line_number,
                            excerpt=_truncate(excerpt, config.max_context_excerpt_chars),
                        )
                    ],
                )
            )
        return fields, [], warnings


class InvoiceStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.invoice
    field_patterns = {
        "invoice_number": [_rx(r"\b(?:invoice|inv)\s*(?:no\.?|number|#)\s*[:#\-]?\s*(?P<value>[A-Z0-9][A-Z0-9\-\/]+)\b")],
        "invoice_date": [_rx(r"\b(?:invoice\s*)?date\s*[:\-]?\s*(?P<value>\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4}|[A-Za-z]+\s+\d{1,2},?\s+\d{4})")],
        "due_date": [_rx(r"\bdue\s+date\s*[:\-]?\s*(?P<value>\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4}|[A-Za-z]+\s+\d{1,2},?\s+\d{4})")],
        "subtotal": [_money_rx("subtotal")],
        "tax": [_money_rx("tax|vat")],
        "total": [_money_rx("total|amount\s+due|balance\s+due")],
    }

    def extract(self, **kwargs: Any) -> tuple[list[ExtractedField], list[dict[str, Any]], list[str]]:
        fields, rows, warnings = super().extract(**kwargs)
        rows.extend(_extract_line_item_rows(kwargs["lines"], kwargs["source_document_index"]))
        return fields, rows, warnings


class ReceiptStrategy(InvoiceStrategy):
    document_class = StructuredExtractionDocumentClass.receipt
    field_patterns = {
        "receipt_number": [_rx(r"\b(?:receipt|rcpt)\s*(?:no\.?|number|#)\s*[:#\-]?\s*(?P<value>[A-Z0-9][A-Z0-9\-\/]+)\b")],
        "transaction_date": [_rx(r"\b(?:date|transaction\s+date)\s*[:\-]?\s*(?P<value>\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4}|[A-Za-z]+\s+\d{1,2},?\s+\d{4})")],
        "merchant": [_rx(r"\b(?:merchant|vendor|seller)\s*[:\-]?\s*(?P<value>.+)$")],
        "subtotal": [_money_rx("subtotal")],
        "tax": [_money_rx("tax|vat")],
        "total": [_money_rx("total|amount\s+paid")],
    }


class BankStatementStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.bank_statement
    field_patterns = {
        "account_name": [_rx(r"\baccount\s+name\s*[:\-]?\s*(?P<value>.+)$")],
        "account_number": [_rx(r"\baccount\s+(?:no\.?|number)\s*[:#\-]?\s*(?P<value>[0-9Xx*\-\s]{5,})")],
        "statement_period": [_rx(r"\b(?:statement\s+period|period)\s*[:\-]?\s*(?P<value>.+)$")],
        "opening_balance": [_money_rx("opening\s+balance")],
        "closing_balance": [_money_rx("closing\s+balance")],
    }

    def extract(self, **kwargs: Any) -> tuple[list[ExtractedField], list[dict[str, Any]], list[str]]:
        fields, rows, warnings = super().extract(**kwargs)
        rows.extend(_extract_transaction_rows(kwargs["lines"], kwargs["source_document_index"]))
        return fields, rows, warnings


class KycDocumentStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.kyc_document
    field_patterns = {
        "full_name": [_rx(r"\b(?:full\s+name|name)\s*[:\-]?\s*(?P<value>[A-Za-z][A-Za-z\s.'-]{2,})$")],
        "date_of_birth": [_rx(r"\b(?:date\s+of\s+birth|dob)\s*[:\-]?\s*(?P<value>\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4})")],
        "national_id": [_rx(r"\b(?:national\s+id|nin|id\s+number)\s*[:#\-]?\s*(?P<value>[A-Z0-9\-]{5,})")],
        "phone_number": [_rx(r"\b(?:phone|mobile|telephone)\s*[:\-]?\s*(?P<value>\+?[0-9][0-9\s()\-]{7,})")],
        "email_address": [_rx(r"\b(?:email|e-mail)\s*[:\-]?\s*(?P<value>[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})")],
        "address": [_rx(r"\b(?:address|residential\s+address)\s*[:\-]?\s*(?P<value>.+)$")],
    }


class IdDocumentStrategy(KycDocumentStrategy):
    document_class = StructuredExtractionDocumentClass.id_document


class ContractStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.contract
    field_patterns = {
        "effective_date": [_rx(r"\b(?:effective\s+date|commencement\s+date)\s*[:\-]?\s*(?P<value>.+)$")],
        "termination_date": [_rx(r"\b(?:termination\s+date|expiry\s+date|expiration\s+date)\s*[:\-]?\s*(?P<value>.+)$")],
        "governing_law": [_rx(r"\bgoverning\s+law\s*[:\-]?\s*(?P<value>.+)$")],
        "party_a": [_rx(r"\b(?:party\s+a|first\s+party|between)\s*[:\-]?\s*(?P<value>.+)$")],
        "party_b": [_rx(r"\b(?:party\s+b|second\s+party|and)\s*[:\-]?\s*(?P<value>.+)$")],
    }

    def extract(self, **kwargs: Any) -> tuple[list[ExtractedField], list[dict[str, Any]], list[str]]:
        fields, rows, warnings = super().extract(**kwargs)
        rows.extend(_extract_clause_rows(kwargs["lines"], kwargs["source_document_index"]))
        return fields, rows, warnings


class LegalRecordStrategy(ContractStrategy):
    document_class = StructuredExtractionDocumentClass.legal_record


class MedicalRecordStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.medical_record
    field_patterns = {
        "patient_name": [_rx(r"\b(?:patient\s+name|name)\s*[:\-]?\s*(?P<value>[A-Za-z][A-Za-z\s.'-]{2,})$")],
        "patient_id": [_rx(r"\b(?:patient\s+id|hospital\s+number|medical\s+record\s+number)\s*[:#\-]?\s*(?P<value>[A-Z0-9\-\/]{3,})")],
        "date_of_birth": [_rx(r"\b(?:date\s+of\s+birth|dob)\s*[:\-]?\s*(?P<value>\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4})")],
        "diagnosis": [_rx(r"\bdiagnosis\s*[:\-]?\s*(?P<value>.+)$")],
        "provider": [_rx(r"\b(?:doctor|physician|provider|consultant)\s*[:\-]?\s*(?P<value>.+)$")],
    }


class ProcurementStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.procurement_document
    field_patterns = {
        "purchase_order_number": [_rx(r"\b(?:purchase\s+order|po)\s*(?:no\.?|number|#)?\s*[:#\-]?\s*(?P<value>[A-Z0-9\-\/]+)")],
        "vendor": [_rx(r"\b(?:vendor|supplier)\s*[:\-]?\s*(?P<value>.+)$")],
        "delivery_date": [_rx(r"\b(?:delivery\s+date|required\s+by)\s*[:\-]?\s*(?P<value>.+)$")],
        "total": [_money_rx("total|grand\s+total|amount")],
    }

    def extract(self, **kwargs: Any) -> tuple[list[ExtractedField], list[dict[str, Any]], list[str]]:
        fields, rows, warnings = super().extract(**kwargs)
        rows.extend(_extract_line_item_rows(kwargs["lines"], kwargs["source_document_index"]))
        return fields, rows, warnings


class InsuranceStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.insurance_document
    field_patterns = {
        "policy_number": [_rx(r"\bpolicy\s*(?:no\.?|number|#)\s*[:#\-]?\s*(?P<value>[A-Z0-9\-\/]+)")],
        "insured_name": [_rx(r"\b(?:insured|policyholder)\s*(?:name)?\s*[:\-]?\s*(?P<value>.+)$")],
        "premium": [_money_rx("premium")],
        "coverage_period": [_rx(r"\b(?:coverage\s+period|policy\s+period)\s*[:\-]?\s*(?P<value>.+)$")],
        "claim_number": [_rx(r"\bclaim\s*(?:no\.?|number|#)\s*[:#\-]?\s*(?P<value>[A-Z0-9\-\/]+)")],
    }


class HRRecordStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.hr_record
    field_patterns = {
        "employee_name": [_rx(r"\b(?:employee\s+name|staff\s+name|name)\s*[:\-]?\s*(?P<value>[A-Za-z][A-Za-z\s.'-]{2,})$")],
        "employee_id": [_rx(r"\b(?:employee\s+id|staff\s+id)\s*[:#\-]?\s*(?P<value>[A-Z0-9\-\/]+)")],
        "department": [_rx(r"\bdepartment\s*[:\-]?\s*(?P<value>.+)$")],
        "job_title": [_rx(r"\b(?:job\s+title|role|position)\s*[:\-]?\s*(?P<value>.+)$")],
        "start_date": [_rx(r"\b(?:start\s+date|employment\s+date|hire\s+date)\s*[:\-]?\s*(?P<value>.+)$")],
    }


class OnboardingStrategy(HRRecordStrategy):
    document_class = StructuredExtractionDocumentClass.onboarding_document


class TicketStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.ticket
    field_patterns = {
        "ticket_id": [_rx(r"\b(?:ticket|case|issue)\s*(?:id|no\.?|number|#)?\s*[:#\-]?\s*(?P<value>[A-Z0-9\-\/]+)")],
        "status": [_rx(r"\bstatus\s*[:\-]?\s*(?P<value>.+)$")],
        "priority": [_rx(r"\bpriority\s*[:\-]?\s*(?P<value>.+)$")],
        "assignee": [_rx(r"\b(?:assignee|assigned\s+to|owner)\s*[:\-]?\s*(?P<value>.+)$")],
        "created_date": [_rx(r"\b(?:created|opened|reported)\s*(?:date)?\s*[:\-]?\s*(?P<value>.+)$")],
    }


class ReportStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.technical_report
    field_patterns = {
        "report_title": [_rx(r"\b(?:report\s+title|title)\s*[:\-]?\s*(?P<value>.+)$")],
        "report_date": [_rx(r"\b(?:report\s+date|date)\s*[:\-]?\s*(?P<value>.+)$")],
        "author": [_rx(r"\b(?:author|prepared\s+by|reported\s+by)\s*[:\-]?\s*(?P<value>.+)$")],
        "summary": [_rx(r"\b(?:summary|executive\s+summary)\s*[:\-]?\s*(?P<value>.+)$")],
    }


class IncidentReportStrategy(ReportStrategy):
    document_class = StructuredExtractionDocumentClass.incident_report
    field_patterns = {
        **ReportStrategy.field_patterns,
        "incident_date": [_rx(r"\bincident\s+date\s*[:\-]?\s*(?P<value>.+)$")],
        "incident_location": [_rx(r"\b(?:incident\s+location|location)\s*[:\-]?\s*(?P<value>.+)$")],
        "severity": [_rx(r"\bseverity\s*[:\-]?\s*(?P<value>.+)$")],
    }


class GenericFormStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.form


class MemoStrategy(BaseRegexStrategy):
    document_class = StructuredExtractionDocumentClass.memo
    field_patterns = {
        "to": [_rx(r"^to\s*[:\-]?\s*(?P<value>.+)$")],
        "from": [_rx(r"^from\s*[:\-]?\s*(?P<value>.+)$")],
        "date": [_rx(r"^date\s*[:\-]?\s*(?P<value>.+)$")],
        "subject": [_rx(r"^(?:subject|re)\s*[:\-]?\s*(?P<value>.+)$")],
    }


# =========================
# Shape enforcement
# =========================


class ShapeEnforcer:
    @staticmethod
    def enforce(
        *,
        result_shape: StructuredExtractionResultShape,
        selected_fields: Sequence[str],
        documents: Sequence[DocumentExtraction],
    ) -> dict[str, Any]:
        base = {
            "contract_version": "v1",
            "feature": FeatureType.structured_extract.value,
            "result_shape": result_shape.value,
            "selected_fields": list(selected_fields),
            "generated_at": _timestamp_iso(),
            "human_review_required": True,
        }

        if result_shape == StructuredExtractionResultShape.key_value_fields:
            base["documents"] = [
                {
                    "source_document_index": doc.source_document_index,
                    "filename": doc.filename,
                    "fields": doc.fields_as_mapping(),
                    "evidence": [evidence.to_dict() for field_item in doc.fields for evidence in field_item.evidence],
                    "warnings": list(doc.warnings),
                }
                for doc in documents
            ]
            return base

        if result_shape == StructuredExtractionResultShape.tables:
            base["documents"] = [
                {
                    "source_document_index": doc.source_document_index,
                    "filename": doc.filename,
                    "tables": [table.to_dict() for table in doc.tables],
                    "warnings": list(doc.warnings),
                }
                for doc in documents
            ]
            return base

        if result_shape == StructuredExtractionResultShape.row_based_records:
            rows: list[dict[str, Any]] = []
            for doc in documents:
                rows.extend([dict(row) for row in doc.row_records])
            base["rows"] = rows
            return base

        if result_shape == StructuredExtractionResultShape.machine_readable:
            base["documents"] = [doc.to_machine_readable() for doc in documents]
            base["aggregate"] = _aggregate_documents(documents)
            return base

        raise ValueError(f"Unsupported result_shape: {result_shape.value}")

    @staticmethod
    def to_tabular_rows(
        *,
        result_shape: StructuredExtractionResultShape,
        selected_fields: Sequence[str],
        documents: Sequence[DocumentExtraction],
    ) -> list[dict[str, Any]]:
        if result_shape == StructuredExtractionResultShape.key_value_fields:
            return [_document_fields_to_row(doc, selected_fields=selected_fields) for doc in documents]

        if result_shape == StructuredExtractionResultShape.tables:
            rows: list[dict[str, Any]] = []
            for doc in documents:
                if not doc.tables:
                    rows.append({"source_document_index": doc.source_document_index, "filename": doc.filename})
                    continue
                for table in doc.tables:
                    for row in table.rows:
                        normalized = {
                            "source_document_index": doc.source_document_index,
                            "filename": doc.filename,
                            "table_index": table.table_index,
                        }
                        normalized.update(row)
                        rows.append(normalized)
            return rows

        if result_shape == StructuredExtractionResultShape.row_based_records:
            rows = []
            for doc in documents:
                rows.extend([dict(row) for row in doc.row_records])
            return rows or [{"source_document_index": 0}]

        rows = []
        for doc in documents:
            row = _document_fields_to_row(doc, selected_fields=selected_fields)
            row["tables_count"] = len(doc.tables)
            row["row_records_count"] = len(doc.row_records)
            row["warnings"] = "; ".join(doc.warnings)
            rows.append(row)
        return rows


# =========================
# Writer layer
# =========================


class StructuredArtifactWriter(Protocol):
    def write(
        self,
        *,
        output_format: StructuredDataOutputFormat,
        payload: Mapping[str, Any],
        rows: Sequence[Mapping[str, Any]],
        stem: str,
    ) -> PersistedArtifact:
        ...


class LocalStructuredArtifactWriter:
    def __init__(self, base_dir: str) -> None:
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)

    def write(
        self,
        *,
        output_format: StructuredDataOutputFormat,
        payload: Mapping[str, Any],
        rows: Sequence[Mapping[str, Any]],
        stem: str,
    ) -> PersistedArtifact:
        if output_format == StructuredDataOutputFormat.json:
            return self._write_json(payload=payload, file_name=f"{stem}.json")
        if output_format == StructuredDataOutputFormat.csv:
            return self._write_csv(rows=rows, file_name=f"{stem}.csv")
        if output_format == StructuredDataOutputFormat.xlsx:
            return self._write_xlsx(payload=payload, rows=rows, file_name=f"{stem}.xlsx")
        raise ValueError(f"Unsupported structured extraction output format: {output_format.value}")

    def _write_json(self, *, payload: Mapping[str, Any], file_name: str) -> PersistedArtifact:
        target = self.base_dir / _safe_filename(file_name)
        target.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return _artifact_from_path(target)

    def _write_csv(self, *, rows: Sequence[Mapping[str, Any]], file_name: str) -> PersistedArtifact:
        target = self.base_dir / _safe_filename(file_name)
        normalized_rows = [_stringify_row(row) for row in (rows or [{"source_document_index": 0}])]
        headers = _headers_for_rows(normalized_rows)
        with target.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            for row in normalized_rows:
                writer.writerow({header: row.get(header, "") for header in headers})
        return _artifact_from_path(target)

    def _write_xlsx(
        self,
        *,
        payload: Mapping[str, Any],
        rows: Sequence[Mapping[str, Any]],
        file_name: str,
    ) -> PersistedArtifact:
        target = self.base_dir / _safe_filename(file_name)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "summary"
        _write_summary_sheet(summary, payload)

        data_sheet = workbook.create_sheet("records")
        normalized_rows = [_stringify_row(row) for row in (rows or [{"source_document_index": 0}])]
        _write_rows_sheet(data_sheet, normalized_rows)

        evidence_sheet = workbook.create_sheet("evidence")
        _write_rows_sheet(evidence_sheet, _evidence_rows_from_payload(payload))

        workbook.save(target)
        return _artifact_from_path(target)


# Backward-compatible alias from the earlier scaffold.
LocalStructuredExtractionArtifactStore = LocalStructuredArtifactWriter


# =========================
# Engine facade
# =========================


class StructuredExtractionEngine:
    def __init__(
        self,
        backend: Optional[StructuredExtractionBackend] = None,
        writer: Optional[StructuredArtifactWriter] = None,
        config: Optional[StructuredExtractionConfig] = None,
    ) -> None:
        self.config = config or StructuredExtractionConfig()
        self.backend = backend or DeterministicStructuredExtractionBackend()
        self.writer = writer or LocalStructuredArtifactWriter(self.config.artifact_base_dir)

    def run(self, request: Union[AnalyzerRequest, Mapping[str, Any]]) -> AnalyzerResponse:
        req = validate_analyzer_request(request)
        self._validate_engine_scope(req)
        payload = req.payload
        assert isinstance(payload, StructuredExtractionRequest)

        documents = _iter_request_documents(req)
        extraction_output = self.backend.extract(
            documents=documents,
            result_shape=payload.result_shape,
            selected_fields=payload.selected_fields,
            document_classes=payload.document_classes,
            config=self.config,
        )

        artifact = self.writer.write(
            output_format=payload.output_format,
            payload=extraction_output.payload,
            rows=extraction_output.tabular_rows,
            stem=f"structured_extract_{payload.result_shape.value}_{_timestamp_slug()}",
        )

        response = AnalyzerResponse(
            action=FeatureType.structured_extract,
            input_format=_expected_response_input_format(req),
            policy=req.policy,
            system_language=req.system_language,
            detected_language=None,
            output_language=None,
            result=build_structured_extraction_file_result(
                filename=artifact.file_name,
                output_format=payload.output_format,
                file_size_mb=artifact.file_size_mb,
                result_shape=payload.result_shape,
                selected_fields=list(payload.selected_fields),
                storage_key=artifact.storage_key,
                download_url=artifact.download_url,
                algorithm_version=self.config.algorithm_version,
            ),
            human_review=HumanReviewRequirement(),
        )
        return validate_analyzer_response(response, request=req)

    def _validate_engine_scope(self, request: AnalyzerRequest) -> None:
        if request.action != FeatureType.structured_extract:
            raise ValueError("StructuredExtractionEngine only handles the structured_extract action.")
        if not isinstance(request.payload, StructuredExtractionRequest):
            raise ValueError("structured_extract requires StructuredExtractionRequest payload.")
        if not isinstance(request.policy, OutputPolicy):
            raise ValueError("structured_extract requires OutputPolicy.")


def run_structured_extraction(request: Union[AnalyzerRequest, Mapping[str, Any]]) -> AnalyzerResponse:
    return StructuredExtractionEngine().run(request)


# =========================
# Extraction helpers
# =========================


def _default_strategies() -> list[DocumentClassExtractionStrategy]:
    return [
        GenericFormStrategy(),
        MemoStrategy(),
        InvoiceStrategy(),
        ReceiptStrategy(),
        BankStatementStrategy(),
        KycDocumentStrategy(),
        IdDocumentStrategy(),
        ContractStrategy(),
        LegalRecordStrategy(),
        MedicalRecordStrategy(),
        ProcurementStrategy(),
        ReportStrategy(),
        IncidentReportStrategy(),
        InsuranceStrategy(),
        HRRecordStrategy(),
        OnboardingStrategy(),
        TicketStrategy(),
    ]


def _iter_request_documents(request: AnalyzerRequest) -> list[DocumentPayload]:
    if isinstance(request.input, DocumentPayload):
        return [request.input]
    if isinstance(request.input, DocumentSetPayload):
        return list(request.input.documents)
    raise ValueError("structured_extract requires DocumentPayload or DocumentSetPayload input.")


def _ensure_document_text(document: DocumentPayload) -> str:
    if document.text and document.text.strip():
        return document.text.strip()

    filename = (document.filename or "").strip()
    if filename:
        path = Path(filename)
        if path.exists():
            extracted_text, _ = extract_text_by_format(path, document.metadata.input_format)
            normalized = extracted_text.strip()
            if normalized:
                return normalized

    raise ValueError(
        "structured_extract requires accessible document text for processing. "
        "Provide DocumentPayload.text or a valid filename reference that can be re-opened."
    )


def _extract_generic_key_values(
    *,
    lines: Sequence[str],
    source_document_index: int,
    selected_fields: Sequence[str],
    config: StructuredExtractionConfig,
) -> list[ExtractedField]:
    fields: list[ExtractedField] = []
    seen: set[str] = set()
    pattern = re.compile(
        r"^\s*(?P<key>[A-Za-z0-9][A-Za-z0-9 /_()#&.,-]{0,80}?)\s*(?:[:\-–—]|\|)\s+(?P<value>.+?)\s*$"
    )
    for line_number, raw_line in enumerate(lines, start=1):
        match = pattern.match(raw_line)
        if not match:
            continue
        key = _clean_key(match.group("key"))
        value = _clean_value(match.group("value"))
        if not key or not value or len(key) > 70 or key.count(" ") > 10:
            continue
        normalized_key = _normalize_field_name(key)
        if selected_fields and normalized_key not in selected_fields:
            continue
        if normalized_key in seen:
            continue
        seen.add(normalized_key)
        fields.append(
            ExtractedField(
                name=key,
                value=value,
                source_document_index=source_document_index,
                confidence=0.88,
                evidence=[
                    SourceEvidence(
                        source_document_index=source_document_index,
                        field_name=key,
                        value=value,
                        line_number=line_number,
                        excerpt=_truncate(raw_line, config.max_context_excerpt_chars),
                    )
                ],
            )
        )
    return fields


def _extract_generic_tables(*, lines: Sequence[str], source_document_index: int) -> list[ExtractedTable]:
    groups = _group_tabular_lines(lines)
    tables: list[ExtractedTable] = []
    for table_number, group in enumerate(groups, start=1):
        header, *body_rows = group
        columns = _split_tabular_line(header)
        if len(columns) < 2:
            continue
        rows: list[dict[str, Any]] = []
        for raw_row in body_rows:
            cells = _split_tabular_line(raw_row)
            if len(cells) != len(columns):
                continue
            rows.append(dict(zip(columns, cells)))
        if rows:
            tables.append(
                ExtractedTable(
                    table_index=table_number,
                    source_document_index=source_document_index,
                    columns=columns,
                    rows=rows,
                )
            )
    return tables


def _extract_line_item_rows(lines: Sequence[str], source_document_index: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    money = r"(?:NGN|₦|USD|\$|EUR|€|GBP|£)?\s?[0-9][0-9,]*(?:\.\d{2})?"
    pattern = re.compile(
        rf"^(?P<description>[A-Za-z][A-Za-z0-9 .,'\-/()]+?)\s+(?P<quantity>\d+(?:\.\d+)?)\s+(?P<unit_price>{money})\s+(?P<amount>{money})$",
        re.IGNORECASE,
    )
    for line in lines:
        match = pattern.match(line.strip())
        if not match:
            continue
        row = {"source_document_index": source_document_index, "record_type": "line_item"}
        row.update({key: _clean_value(value) for key, value in match.groupdict().items()})
        rows.append(row)
    return rows


def _extract_transaction_rows(lines: Sequence[str], source_document_index: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    date = r"\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4}"
    money = r"(?:NGN|₦|USD|\$|EUR|€|GBP|£)?\s?[0-9][0-9,]*(?:\.\d{2})?"
    pattern = re.compile(
        rf"^(?P<date>{date})\s+(?P<description>.+?)\s+(?P<amount>-?{money})(?:\s+(?P<balance>{money}))?$",
        re.IGNORECASE,
    )
    for line in lines:
        match = pattern.match(line.strip())
        if not match:
            continue
        row = {"source_document_index": source_document_index, "record_type": "transaction"}
        row.update({key: _clean_value(value or "") for key, value in match.groupdict().items()})
        rows.append(row)
    return rows


def _extract_clause_rows(lines: Sequence[str], source_document_index: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    pattern = re.compile(r"^(?P<section>\d+(?:\.\d+)*)\s*[.)]?\s+(?P<title>[A-Z][A-Za-z0-9 ,&/()\-]{2,80})")
    for line in lines:
        match = pattern.match(line.strip())
        if not match:
            continue
        rows.append(
            {
                "source_document_index": source_document_index,
                "record_type": "clause",
                "section": match.group("section"),
                "title": match.group("title").strip(),
            }
        )
    return rows


def _derive_row_records(
    *,
    source_document_index: int,
    fields: Sequence[ExtractedField],
    tables: Sequence[ExtractedTable],
    class_rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = [dict(row) for row in class_rows]
    if rows:
        return rows
    if tables:
        for table in tables:
            for row in table.rows:
                normalized = {
                    "source_document_index": source_document_index,
                    "record_type": "table_row",
                    "table_index": table.table_index,
                }
                normalized.update(dict(row))
                rows.append(normalized)
        return rows
    if fields:
        row = {"source_document_index": source_document_index, "record_type": "field_set"}
        for field_item in fields:
            row[field_item.name] = field_item.value
        return [row]
    return [{"source_document_index": source_document_index, "record_type": "empty"}]


def _merge_fields(
    fields: Sequence[ExtractedField],
    *,
    selected_fields: Sequence[str],
    include_empty_selected_fields: bool,
    source_document_index: int,
) -> list[ExtractedField]:
    by_normalized: dict[str, ExtractedField] = {}
    for field_item in fields:
        normalized = _normalize_field_name(field_item.name)
        if selected_fields and normalized not in selected_fields:
            continue
        existing = by_normalized.get(normalized)
        if existing is None or field_item.confidence > existing.confidence:
            by_normalized[normalized] = field_item

    if include_empty_selected_fields:
        for raw_selected in selected_fields:
            if raw_selected not in by_normalized:
                by_normalized[raw_selected] = ExtractedField(
                    name=raw_selected,
                    value="",
                    source_document_index=source_document_index,
                    confidence=0.0,
                    evidence=[],
                )

    return sorted(by_normalized.values(), key=lambda item: _normalize_field_name(item.name))


def _first_regex_match(lines: Sequence[str], patterns: Sequence[re.Pattern[str]]) -> Optional[tuple[str, int, str]]:
    for line_number, line in enumerate(lines, start=1):
        for pattern in patterns:
            match = pattern.search(line)
            if match:
                value = _clean_value(match.group("value"))
                if value:
                    return value, line_number, line
    return None


# =========================
# Generic utilities
# =========================

def _iter_lines(text: str) -> Iterable[str]:
    for line in text.splitlines():
        normalized = line.strip()
        if normalized:
            yield normalized


def _normalize_selected_fields(fields: Sequence[str]) -> list[str]:
    output: list[str] = []
    for item in fields:
        normalized = _normalize_field_name(item)
        if normalized and normalized not in output:
            output.append(normalized)
    return output


def _normalize_field_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def _clean_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().strip(":-–—|"))


def _clean_value(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip().strip("|"))


def _truncate(value: Optional[str], max_chars: int) -> Optional[str]:
    if value is None:
        return None
    text = str(value)
    if len(text) <= max_chars:
        return text
    return f"{text[: max_chars - 1]}…"


def _split_tabular_line(line: str) -> list[str]:
    stripped = line.strip()
    if "|" in stripped:
        return [_clean_value(cell) for cell in stripped.strip("|").split("|")]
    if "\t" in stripped:
        return [_clean_value(cell) for cell in stripped.split("\t")]
    if ";" in stripped:
        cells = [_clean_value(cell) for cell in stripped.split(";")]
        if len(cells) >= 2:
            return cells
    if "," in stripped:
        cells = [_clean_value(cell) for cell in stripped.split(",")]
        if len(cells) >= 2:
            return cells
    return [stripped]


def _looks_tabular(line: str) -> bool:
    return "|" in line or "\t" in line or line.count(",") >= 1 or line.count(";") >= 1


def _group_tabular_lines(lines: Sequence[str]) -> list[list[str]]:
    groups: list[list[str]] = []
    current: list[str] = []
    for line in lines:
        if _looks_tabular(line):
            current.append(line)
            continue
        if len(current) >= 2:
            groups.append(current)
        current = []
    if len(current) >= 2:
        groups.append(current)
    return groups


def _aggregate_documents(documents: Sequence[DocumentExtraction]) -> dict[str, Any]:
    return {
        "document_count": len(documents),
        "field_count": sum(len(doc.fields) for doc in documents),
        "table_count": sum(len(doc.tables) for doc in documents),
        "row_record_count": sum(len(doc.row_records) for doc in documents),
        "warnings_count": sum(len(doc.warnings) for doc in documents),
    }


def _document_fields_to_row(doc: DocumentExtraction, *, selected_fields: Sequence[str]) -> dict[str, Any]:
    row: dict[str, Any] = {
        "source_document_index": doc.source_document_index,
        "filename": doc.filename or "",
        "input_format": doc.input_format,
        "ocr_used": doc.ocr_used,
    }
    fields_map = doc.fields_as_mapping()
    if selected_fields:
        normalized_to_original = {_normalize_field_name(key): key for key in fields_map.keys()}
        for selected in selected_fields:
            actual_key = normalized_to_original.get(_normalize_field_name(selected), selected)
            row[selected] = fields_map.get(actual_key, "")
    else:
        row.update(fields_map)
    return row


def _headers_for_rows(rows: Sequence[Mapping[str, Any]]) -> list[str]:
    preferred = ["source_document_index", "filename", "input_format", "ocr_used", "record_type", "table_index"]
    headers: list[str] = []
    for item in preferred:
        if any(item in row for row in rows):
            headers.append(item)
    for row in rows:
        for key in row.keys():
            key_str = str(key)
            if key_str not in headers:
                headers.append(key_str)
    return headers or ["source_document_index"]


def _stringify_row(row: Mapping[str, Any]) -> dict[str, str]:
    output: dict[str, str] = {}
    for key, value in row.items():
        if isinstance(value, (dict, list)):
            output[str(key)] = json.dumps(value, ensure_ascii=False, default=str)
        elif value is None:
            output[str(key)] = ""
        else:
            output[str(key)] = str(value)
    return output


def _write_summary_sheet(sheet: Worksheet, payload: Mapping[str, Any]) -> None:
    rows = [
        ("contract_version", payload.get("contract_version", "")),
        ("feature", payload.get("feature", "")),
        ("result_shape", payload.get("result_shape", "")),
        ("generated_at", payload.get("generated_at", "")),
        ("human_review_required", payload.get("human_review_required", "")),
        ("selected_fields", ", ".join(payload.get("selected_fields", []) or [])),
    ]
    if isinstance(payload.get("aggregate"), Mapping):
        for key, value in payload["aggregate"].items():
            rows.append((f"aggregate.{key}", value))
    sheet.append(["property", "value"])
    for row in rows:
        sheet.append(list(row))


def _write_rows_sheet(sheet: Worksheet, rows: Sequence[Mapping[str, Any]]) -> None:
    normalized_rows = [_stringify_row(row) for row in (rows or [{"source_document_index": 0}])]
    headers = _headers_for_rows(normalized_rows)
    sheet.append(headers)
    for row in normalized_rows:
        sheet.append([row.get(header, "") for header in headers])


def _evidence_rows_from_payload(payload: Mapping[str, Any]) -> list[dict[str, Any]]:
    evidence_rows: list[dict[str, Any]] = []
    documents = payload.get("documents")
    if not isinstance(documents, list):
        return evidence_rows
    for document in documents:
        if not isinstance(document, Mapping):
            continue
        if isinstance(document.get("evidence"), list):
            evidence_rows.extend([dict(item) for item in document["evidence"] if isinstance(item, Mapping)])
        fields = document.get("fields")
        if isinstance(fields, list):
            for field_item in fields:
                if not isinstance(field_item, Mapping):
                    continue
                for evidence in field_item.get("evidence", []) or []:
                    if isinstance(evidence, Mapping):
                        evidence_rows.append(dict(evidence))
    return evidence_rows or [{"source_document_index": "", "field_name": "", "value": "", "line_number": "", "excerpt": ""}]


def _timestamp_slug() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _timestamp_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _artifact_from_path(path: Path) -> PersistedArtifact:
    size_mb = round(path.stat().st_size / (1024 * 1024), 4)
    return PersistedArtifact(
        file_name=path.name,
        file_extension=path.suffix.lstrip("."),
        file_size_mb=size_mb,
        file_path=str(path),
        storage_key=str(path),
        download_url=None,
    )


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._")
    if not cleaned:
        raise ValueError("Artifact filename cannot be empty.")
    return cleaned


def _expected_response_input_format(request: AnalyzerRequest) -> Union[str, Any]:
    if isinstance(request.input, DocumentSetPayload):
        return "document_set"
    assert isinstance(request.input, DocumentPayload)
    return request.input.metadata.input_format


__all__ = [
    "STRUCTURED_EXTRACTION_RULES",
    "PersistedArtifact",
    "StructuredExtractionConfig",
    "SourceEvidence",
    "ExtractedField",
    "ExtractedTable",
    "DocumentExtraction",
    "StructuredExtractionOutput",
    "DocumentClassExtractionStrategy",
    "StructuredExtractionBackend",
    "DeterministicStructuredExtractionBackend",
    "ShapeEnforcer",
    "StructuredArtifactWriter",
    "LocalStructuredArtifactWriter",
    "LocalStructuredExtractionArtifactStore",
    "StructuredExtractionEngine",
    "run_structured_extraction",
]
